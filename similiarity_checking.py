import pandas as pd
import numpy as np
from openpyxl import load_workbook
def jaro_distance(s1, s2):
    if (s1 == s2):
        return 1.0
    len1 = len(s1)
    len2 = len(s2)
    max_dist = floor(max(len1, len2) / 2) - 1
    match = 0
    hash_s1 = [0] * len(s1)
    hash_s2 = [0] * len(s2)
    for i in range(len1):
        for j in range(max(0, i - max_dist), min(len2, i + max_dist + 1)):
            if (s1[i] == s2[j] and hash_s2[j] == 0):
                hash_s1[i] = 1
                hash_s2[j] = 1
                match += 1
                break
    if (match == 0):
        return 0.0
    t = 0
    point = 0
    for i in range(len1):
        if (hash_s1[i]):

            while (hash_s2[point] == 0):
                point += 1
            if (s1[i] != s2[point]):
                t += 1
            point += 1
    t = t // 2
    return (match / len1 + match / len2 + (match - t) / match) / 3.0

if __name__ == "__main__":
    file_main=r
    file_tobechecked=r
    a = pd.read_excel(file_main,header = 0)
    b = pd.read_excel(file_tobechecked,header = 0)
    similarity=[]
    company=[]
    banned_word = ["Mrs","Miss","Mr","Sir","Madam"]
    for i in a.iloc[:,1]:
        save = []
        for j in b.iloc[:,1]:
            tempory_1 =i.lower()
            tempory_2 = j.lower()
            for ban in ban_word:
                tempory_1= tempory_1.replace(ban,"")
                tempory_2= tempory_2.replace(ban,"")
            save.append(jaro_distance(i.lower(), j.lower()))
        company.append(dataset.iloc[save.index(np.max(save))].loc["Name"])
        similarity.append(np.max(save))
    a["Company"] = company
    a["Similarity"] = similarity
    wb=load_workbook(file_main)
    ws=wb[wb.sheetnames[0]]
    for i in range(len(a["Company"])):
        ws["B"+str(2+i)] = a.iloc[i].loc["Company"]
        ws["C"+str(2+i)].value = a.iloc[i].loc["similarity"]
        ws["C"+str(2+i)].number_format = "0.00%"
    file_saveto = file_main
    wb.save(file_main)
